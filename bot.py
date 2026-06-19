import os
import logging
import asyncio
from openpyxl import load_workbook
import xlrd
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, ContextTypes, filters
)
from telegram.error import BadRequest
import anthropic
from pypdf import PdfReader
from docx import Document
import io

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")

anthropic_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

# ───────────────────────────────────────────
# ПРОМПТЫ ДЕ-БРИФИНГА
# ───────────────────────────────────────────

SYSTEM_PROMPT_PROJECT = """Ты - старший стратег digital-агентства Wunder Digital. Проведи де-брифинг клиентского брифа по методологии агентства.

ВАЖНО ПО СТИЛЮ: Пиши лаконично. Максимум 2 предложения на каждый блок. Никакой воды и повторений.

ТИП: ПРОЕКТ. Фокус - на бизнес-задаче, бюджете и медиа-миксе. Блоки 6 и 7 менее критичны.

Проанализируй бриф по 7 блокам:
1. Бизнес-задача - какую бизнес-задачу решает клиент. ВАЖНО: если задача сформулирована только как "нужен охват" - это отсутствует, необходимо докопаться до бизнес-причины.
2. Маркетинговая задача - рост знания, формирование спроса, лидогенерация, переключение с конкурентов. Приоритетная задача и этап воронки.
3. Измерение успеха - KPI и метод оценки: продажи, Brand Lift, лиды, доля рынка.
4. Бюджет и рамки - диапазон бюджета, ограничения. Отсутствие бюджета = критический риск.
5. Архитектура запуска и медиа-микс - полный медиа-микс, роль digital, барьеры восприятия бренда.
6. Методология выбора - для проекта менее критично.
7. Критерии победы - для проекта менее критично.

Используй следующий формат ответа СТРОГО (Markdown + эмодзи):

*📋 ДЕ-БРИФ — WUNDER DIGITAL*
_Тип: Проект_

*1️⃣ Бизнес-задача* — [✅ OK / ⚠️ РИСК / ❌ ОТСУТСТВУЕТ]
[резюме — 1-2 предложения]
_❔ Необходимо уточнить:_ [коротко, или "—"]

*2️⃣ Маркетинговая задача* — [✅ OK / ⚠️ РИСК / ❌ ОТСУТСТВУЕТ]
[резюме — 1-2 предложения]
_❔ Необходимо уточнить:_ [...]

*3️⃣ Измерение успеха* — [✅ OK / ⚠️ РИСК / ❌ ОТСУТСТВУЕТ]
[резюме — 1-2 предложения]
_❔ Необходимо уточнить:_ [...]

*4️⃣ Бюджет и рамки* — [✅ OK / ⚠️ РИСК / ❌ ОТСУТСТВУЕТ]
[резюме — 1-2 предложения]
_❔ Необходимо уточнить:_ [...]

*5️⃣ Архитектура запуска* — [✅ OK / ⚠️ РИСК / ❌ ОТСУТСТВУЕТ]
[резюме — 1-2 предложения]
_❔ Необходимо уточнить:_ [...]

*6️⃣ Методология выбора* — [✅ OK / ⚠️ РИСК / 🔵 НЕ ПРИМЕНИМО]
[резюме — 1 предложение]

*7️⃣ Критерии победы* — [✅ OK / ⚠️ РИСК / 🔵 НЕ ПРИМЕНИМО]
[резюме — 1 предложение]

➖➖➖➖➖➖➖➖➖➖

*🚨 Риски*
• [риск — одна строка]

*❓ Вопросы для доуточнения*
• [вопрос — одна строка]

*🚀 Готовность команды*
[1-2 предложения]"""

SYSTEM_PROMPT_TENDER = """Ты - старший стратег digital-агентства Wunder Digital. Проведи де-брифинг тендерного брифа по методологии агентства.

ВАЖНО ПО СТИЛЮ: Пиши лаконично. Максимум 2 предложения на каждый блок. Никакой воды и повторений.

ТИП: ТЕНДЕР. Все 7 блоков критичны. Особый фокус на блоках 6 и 7 - без понимания методологии выбора и критериев победы участие в тендере = слепой риск.

Проанализируй бриф по 7 блокам:
1. Бизнес-задача - какую бизнес-задачу решает клиент. Дополнительно: почему тендер сейчас? Есть ли действующий подрядчик? Чем доволен/недоволен?
2. Маркетинговая задача - рост знания, формирование спроса, лидогенерация, переключение с конкурентов. Приоритетная задача и этап воронки.
3. Измерение успеха - KPI и метод оценки. Отсутствие = риск.
4. Бюджет и рамки - диапазон бюджета, ограничения. Отсутствие = критический риск.
5. Архитектура запуска и медиа-микс - полный медиа-микс, роль digital, барьеры, ограничения по каналам.
6. Методология выбора победителя - КРИТИЧНО. Балльная система? Защита? Кто решает? Кто влияет? Отсутствие = критический риск, сигнал FO.
7. Критерии победы - КРИТИЧНО. Что делает предложение сильным? Отсутствие = критический риск, сигнал FO.

Используй следующий формат ответа СТРОГО (Markdown + эмодзи):

*🏆 ДЕ-БРИФ — WUNDER DIGITAL*
_Тип: Тендер_

*1️⃣ Бизнес-задача* — [✅ OK / ⚠️ РИСК / ❌ ОТСУТСТВУЕТ]
[резюме — 1-2 предложения]
_❔ Необходимо уточнить:_ [коротко, или "—"]

*2️⃣ Маркетинговая задача* — [✅ OK / ⚠️ РИСК / ❌ ОТСУТСТВУЕТ]
[резюме — 1-2 предложения]
_❔ Необходимо уточнить:_ [...]

*3️⃣ Измерение успеха* — [✅ OK / ⚠️ РИСК / ❌ ОТСУТСТВУЕТ]
[резюме — 1-2 предложения]
_❔ Необходимо уточнить:_ [...]

*4️⃣ Бюджет и рамки* — [✅ OK / ⚠️ РИСК / ❌ ОТСУТСТВУЕТ]
[резюме — 1-2 предложения]
_❔ Необходимо уточнить:_ [...]

*5️⃣ Архитектура запуска* — [✅ OK / ⚠️ РИСК / ❌ ОТСУТСТВУЕТ]
[резюме — 1-2 предложения]
_❔ Необходимо уточнить:_ [...]

*6️⃣ Методология выбора* — [✅ OK / ⚠️ РИСК / ❌ ОТСУТСТВУЕТ]
[резюме — если ОТСУТСТВУЕТ: ⚠️ Требует доуточнения через FO до старта]
_❔ Необходимо уточнить:_ [...]

*7️⃣ Критерии победы* — [✅ OK / ⚠️ РИСК / ❌ ОТСУТСТВУЕТ]
[резюме — если ОТСУТСТВУЕТ: ⚠️ Требует доуточнения через FO до старта]
_❔ Необходимо уточнить:_ [...]

➖➖➖➖➖➖➖➖➖➖

*🚨 Риски тендера*
• [риск — одна строка]

*❓ Вопросы для доуточнения*
• [вопрос — одна строка]

*🚀 Готовность команды*
[1-2 предложения]"""

# ───────────────────────────────────────────
# ПРОМПТ АУДИТА МЕДИАПЛАНА
# ───────────────────────────────────────────

SYSTEM_PROMPT_AUDIT = """Ты — старший медиапланер digital-агентства Wunder Digital. Проведи аудит медиаплана по методологии агентства.

ВАЖНО ПО СТИЛЮ: Пиши лаконично. Максимум 2 предложения на каждое замечание. Никакой воды.

КРИТИЧЕСКИ ВАЖНЫЕ ПРАВИЛА:
1. Данные переданы в формате "Заголовок: Значение". Читай внимательно.
2. НИКОГДА не указывай на отсутствие данных если они присутствуют в файле.
3. Если данные есть, но сложно интерпретировать — пиши ВНИМАНИЕ, не КРИТИЧНО.
4. Если поле есть хотя бы в части строк — не считай его отсутствующим глобально.
5. Перед "поле отсутствует" — перечитай данные ещё раз.
6. Скрытые строки и столбцы не проверяются.
7. При неуверенности — пиши "недостаточно данных для оценки".

Проверь по 5 блокам:

БЛОК 1 — СТРУКТУРА И ФОРМАТ
- КРИТИЧНО: Поля отсутствующие полностью ни в одной строке (площадка, формат, период, ЦА, ГЕО, показы/клики/охват, CPM/CPC/CPV, бюджет).
- ВНИМАНИЕ: Разные написания одной площадки или формата.
- ЗАМЕЧАНИЕ: Неединый формат дат и чисел.

БЛОК 2 — БЮДЖЕТ И ФИНАНСЫ
- КРИТИЧНО: Итог не совпадает с суммой строк (если оба значения есть).
- КРИТИЧНО: НДС неединообразен. Казахстан — НДС 16% без исключений.
- ВНИМАНИЕ: Несколько валют без курса конвертации.
- ВНИМАНИЕ: Модель закупки не соответствует задаче.

БЛОК 3 — МЕДИАПОКАЗАТЕЛИ
- КРИТИЧНО: Явные математические ошибки (показы × CTR ≠ клики если оба значения есть).
- ВНИМАНИЕ: CTR баннера > 3% — подозрительно. CPM/CPC вне рыночных норм.
- ВНИМАНИЕ: Охват превышает реалистичный размер ЦА.

БЛОК 4 — ТАРГЕТИНГИ И АУДИТОРИЯ
- КРИТИЧНО: ГЕО отсутствует полностью во всех строках.
- ВНИМАНИЕ: Разные ГЕО без обоснования. Возраст/пол ЦА не указан нигде.

БЛОК 5 — СРОКИ И ПЕРИОДЫ
- КРИТИЧНО: Дата окончания раньше даты начала.
- ВНИМАНИЕ: Пересечения или разрывы между этапами. Даты вне планового периода.

КЛАССИФИКАЦИЯ:
- КРИТИЧНО — явная ошибка с доказательством из данных. Исправить до утверждения.
- ВНИМАНИЕ — влияет на качество. Рекомендуется исправить.
- ЗАМЕЧАНИЕ — рекомендация, не блокирует утверждение.

Формат ответа (Markdown + эмодзи):

*🔍 АУДИТ МЕДИАПЛАНА — WUNDER DIGITAL*

*📊 Сводка*
🔴 Критично: [N]
🟡 Внимание: [N]
🔵 Замечание: [N]
_Вывод:_ [можно утверждать / требуются исправления]

➖➖➖➖➖➖➖➖➖➖

*1️⃣ Структура и формат*
[если нет: ✅ Замечаний нет]
🔴 *КРИТИЧНО — [заголовок]*
Описание: [1 предложение со ссылкой на данные]
Решение: [1 предложение]

🟡 *ВНИМАНИЕ — [заголовок]*
Описание: [1 предложение]
Решение: [1 предложение]

*2️⃣ Бюджет и финансы*
[аналогично]

*3️⃣ Медиапоказатели*
[аналогично]

*4️⃣ Таргетинги и аудитория*
[аналогично]

*5️⃣ Сроки и периоды*
[аналогично]

➖➖➖➖➖➖➖➖➖➖

*✅ Финальный чеклист*
Структура: [OK / требует правок]
Бюджет: [OK / требует правок]
Метрики: [OK / требует правок]
Аудитория: [OK / требует правок]
Сроки: [OK / требует правок]

*🔁 Второй проход завершён.*"""

# ───────────────────────────────────────────
# СОСТОЯНИЯ ПОЛЬЗОВАТЕЛЕЙ
# ───────────────────────────────────────────

user_states = {}

# ───────────────────────────────────────────
# УТИЛИТЫ
# ───────────────────────────────────────────

async def safe_edit(msg, text):
    try:
        await msg.edit_text(text, parse_mode='Markdown')
    except BadRequest:
        try:
            await msg.edit_text(text)
        except Exception:
            pass

async def safe_delete(msg):
    try:
        await msg.delete()
    except Exception:
        pass

async def send_long_message(update, text):
    max_len = 4000
    if len(text) <= max_len:
        try:
            await update.message.reply_text(text, parse_mode='Markdown')
        except Exception:
            await update.message.reply_text(text)
        return
    parts = []
    while text:
        if len(text) <= max_len:
            parts.append(text)
            break
        split_at = text.rfind('\n', 0, max_len)
        if split_at == -1:
            split_at = max_len
        parts.append(text[:split_at])
        text = text[split_at:].lstrip('\n')
    for part in parts:
        try:
            await update.message.reply_text(part, parse_mode='Markdown')
        except Exception:
            await update.message.reply_text(part)

def extract_excel_content(file_bytes: bytes, filename: str) -> str:
    name = filename.lower()
    result = []
    try:
        if name.endswith('.xls'):
            wb = xlrd.open_workbook(file_contents=file_bytes)
            for sheet in wb.sheets():
                rows_raw = []
                for rx in range(sheet.nrows):
                    row = [str(sheet.cell_value(rx, cx)).strip() for cx in range(sheet.ncols)]
                    if any(v for v in row):
                        rows_raw.append(row)
                if not rows_raw:
                    continue
                result.append(f"=== Вкладка: {sheet.name} ===")
                headers = rows_raw[0]
                result.append("Заголовки: " + " | ".join(h for h in headers if h))
                result.append("")
                for row in rows_raw[1:]:
                    labeled = []
                    for i, cell in enumerate(row):
                        if cell:
                            header = headers[i] if i < len(headers) and headers[i] else f"Столбец_{i+1}"
                            labeled.append(f"{header}: {cell}")
                    if labeled:
                        result.append("• " + " | ".join(labeled))
        else:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_raw = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c).strip() if c is not None else "" for c in row]
                    if any(c for c in cells):
                        rows_raw.append(cells)
                if not rows_raw:
                    continue
                result.append(f"=== Вкладка: {sheet_name} ===")
                headers = rows_raw[0]
                result.append("Заголовки: " + " | ".join(h for h in headers if h))
                result.append("")
                for row in rows_raw[1:]:
                    labeled = []
                    for i, cell in enumerate(row):
                        if cell:
                            header = headers[i] if i < len(headers) and headers[i] else f"Столбец_{i+1}"
                            labeled.append(f"{header}: {cell}")
                    if labeled:
                        result.append("• " + " | ".join(labeled))
            wb.close()
        return "\n".join(result) if result else ""
    except Exception as e:
        raise ValueError(f"Не удалось прочитать Excel файл: {e}")

# ───────────────────────────────────────────
# КОМАНДЫ БОТА
# ───────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 *Привет! Я бот Wunder Digital.*\n\n"
        "Что умею:\n"
        "📋 /debrief — анализ клиентского брифа по 7 блокам\n"
        "🔍 /audit — аудит медиаплана на ошибки\n"
        "❓ /help — справка",
        parse_mode='Markdown'
    )

async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "*Команды бота:*\n\n"
        "📋 */debrief* — де-брифинг клиентского брифа\n"
        "Выбери тип (Проект / Тендер), отправь бриф текстом или файлом (PDF, DOCX, TXT)\n\n"
        "🔍 */audit* — аудит медиаплана\n"
        "Отправь файл медиаплана (XLSX, XLS) после команды\n\n"
        "*Чем отличается де-бриф:*\n"
        "Проект — фокус на бизнес-задаче, бюджете, медиа-миксе\n"
        "Тендер — все 7 блоков критичны, особый акцент на методологии выбора и критериях победы",
        parse_mode='Markdown'
    )

async def debrief_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    keyboard = [[
        InlineKeyboardButton("📋 Проект", callback_data="type_project"),
        InlineKeyboardButton("🏆 Тендер", callback_data="type_tender"),
    ]]
    await update.message.reply_text(
        "*Де-бриф запущен.* Выбери тип проекта:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='Markdown'
    )
    user_states[user_id] = {'step': 'choose_type'}

async def audit_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    user_states[user_id] = {'step': 'waiting_mediaplan'}
    await update.message.reply_text(
        "🔍 *Аудит медиаплана запущен.*\n\n"
        "📎 Отправь файл медиаплана в формате XLSX или XLS.\n\n"
        "⚠️ _Скрытые строки, столбцы и вкладки не проверяются._",
        parse_mode='Markdown'
    )

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    if query.data in ('type_project', 'type_tender'):
        project_type = 'Проект' if query.data == 'type_project' else 'Тендер'
        user_states[user_id] = {'step': 'waiting_brief', 'type': project_type}
        hint = (
            "📋 Фокус: бизнес-задача, бюджет, медиа-микс."
            if project_type == 'Проект'
            else "🏆 Фокус: все 7 блоков, особенно методология выбора и критерии победы."
        )
        await query.edit_message_text(
            f"*Тип:* {project_type}\n_{hint}_\n\n📎 Отправь бриф — текстом или файлом (PDF, DOCX, TXT).",
            parse_mode='Markdown'
        )

# ───────────────────────────────────────────
# ОБРАБОТКА ФАЙЛОВ
# ───────────────────────────────────────────

async def extract_text_from_file(file_bytes: bytes, filename: str) -> str:
    name = filename.lower()
    if name.endswith('.pdf'):
        reader = PdfReader(io.BytesIO(file_bytes))
        return "".join(page.extract_text() or "" for page in reader.pages)
    elif name.endswith('.docx') or name.endswith('.doc'):
        doc = Document(io.BytesIO(file_bytes))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    elif name.endswith('.txt'):
        return file_bytes.decode('utf-8', errors='ignore')
    else:
        raise ValueError(f"Формат не поддерживается: {filename}")

async def analyze_brief(brief_text: str, project_type: str) -> str:
    system = SYSTEM_PROMPT_TENDER if project_type == 'Тендер' else SYSTEM_PROMPT_PROJECT
    response = anthropic_client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=8192,
        system=system,
        messages=[{"role": "user", "content": f"Бриф клиента:\n\n{brief_text}"}]
    )
    return response.content[0].text

async def analyze_mediaplan(mp_text: str) -> str:
    response = anthropic_client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=8192,
        system=SYSTEM_PROMPT_AUDIT,
        messages=[{"role": "user", "content": f"Медиаплан для аудита:\n\n{mp_text[:30000]}"}]
    )
    return response.content[0].text

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    state = user_states.get(user_id, {})
    if state.get('step') != 'waiting_brief':
        await update.message.reply_text(
            "Напиши /debrief чтобы начать анализ брифа или /audit для аудита медиаплана."
        )
        return
    brief_text = update.message.text
    if len(brief_text) < 50:
        await update.message.reply_text("⚠️ Текст слишком короткий. Отправь полный текст брифа.")
        return
    project_type = state.get('type', 'Проект')
    thinking_msg = await update.message.reply_text("⏳ Анализирую бриф по 7 блокам...")
    try:
        result = await analyze_brief(brief_text, project_type)
        await safe_delete(thinking_msg)
        await send_long_message(update, result)
        user_states.pop(user_id, None)
    except Exception as e:
        logger.error(f"Error analyzing text: {e}")
        await safe_edit(thinking_msg, "❌ Ошибка при анализе. Попробуй снова — /debrief")

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    state = user_states.get(user_id, {})
    step = state.get('step')

    if step not in ('waiting_brief', 'waiting_mediaplan'):
        await update.message.reply_text(
            "Напиши /debrief для анализа брифа или /audit для аудита медиаплана."
        )
        return

    doc = update.message.document
    filename = doc.file_name or ''
    name = filename.lower()

    if doc.file_size > 20 * 1024 * 1024:
        await update.message.reply_text("⚠️ Файл слишком большой. Максимум 20 МБ.")
        return

    if step == 'waiting_mediaplan':
        if not any(name.endswith(ext) for ext in ['.xlsx', '.xls']):
            await update.message.reply_text("⚠️ Для аудита нужен Excel файл (XLSX или XLS).")
            return
        thinking_msg = await update.message.reply_text(f"📊 Читаю медиаплан {filename}...")
        try:
            file = await context.bot.get_file(doc.file_id)
            file_bytes = await file.download_as_bytearray()
            mp_text = extract_excel_content(bytes(file_bytes), filename)
            if not mp_text.strip():
                await safe_edit(thinking_msg, "❌ Не удалось прочитать данные из файла.")
                return
            await safe_edit(thinking_msg, "🔍 Провожу аудит по 5 блокам...")
            result = await analyze_mediaplan(mp_text)
            await safe_delete(thinking_msg)
            await send_long_message(update, result)
            user_states.pop(user_id, None)
        except Exception as e:
            logger.error(f"Error processing mediaplan: {e}")
            await safe_edit(thinking_msg, f"❌ Ошибка: {str(e)[:100]}. Попробуй снова — /audit")
        return

    if not any(name.endswith(ext) for ext in ['.pdf', '.docx', '.doc', '.txt']):
        await update.message.reply_text("⚠️ Формат не поддерживается. Отправь PDF, DOCX или TXT.")
        return
    thinking_msg = await update.message.reply_text(f"📄 Читаю файл {filename}...")
    try:
        file = await context.bot.get_file(doc.file_id)
        file_bytes = await file.download_as_bytearray()
        brief_text = await extract_text_from_file(bytes(file_bytes), filename)
        if not brief_text.strip():
            await safe_edit(thinking_msg, "❌ Не удалось извлечь текст из файла.")
            return
        project_type = state.get('type', 'Проект')
        await safe_edit(thinking_msg, "⏳ Анализирую бриф по 7 блокам...")
        result = await analyze_brief(brief_text, project_type)
        await safe_delete(thinking_msg)
        await send_long_message(update, result)
        user_states.pop(user_id, None)
    except Exception as e:
        logger.error(f"Error processing document: {e}")
        await safe_edit(thinking_msg, f"❌ Ошибка: {str(e)[:100]}. Попробуй снова — /debrief")

# ───────────────────────────────────────────
# ЗАПУСК
# ───────────────────────────────────────────

async def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("debrief", debrief_start))
    app.add_handler(CommandHandler("audit", audit_start))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    logger.info("Bot started")
    async with app:
        await app.start()
        await app.updater.start_polling(drop_pending_updates=True)
        await asyncio.Event().wait()

if __name__ == "__main__":
    asyncio.run(main())
